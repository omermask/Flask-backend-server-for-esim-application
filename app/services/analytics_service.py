from __future__ import annotations

import csv
import logging
from datetime import datetime, timedelta, timezone
from io import BytesIO, StringIO

from fpdf import FPDF
from sqlalchemy import func
from sqlalchemy.orm import joinedload

from app.core.database import get_session
from app.core.errors import AppError, ErrorCode
from app.models.audit import AuditLog
from app.models.order import Order
from app.models.payment import Payment
from app.models.plan import Plan
from app.models.user import User

logger = logging.getLogger("esim-ego")

_REPORT_TYPES = {"users", "orders", "payments", "financial"}
_EXPORT_FORMATS = {"csv", "pdf", "xlsx"}


class AnalyticsService:

    @staticmethod
    def get_dashboard_stats() -> dict:
        now = datetime.now(timezone.utc)
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        week_start = today_start - timedelta(days=today_start.weekday())
        month_start = today_start.replace(day=1)

        with get_session() as session:
            # Users
            total_users = session.query(func.count(User.id)).scalar() or 0
            active_users = (
                session.query(func.count(User.id))
                .filter(User.is_active.is_(True))
                .scalar()
                or 0
            )
            new_users_today = (
                session.query(func.count(User.id))
                .filter(User.created_at >= today_start)
                .scalar()
                or 0
            )
            new_users_week = (
                session.query(func.count(User.id))
                .filter(User.created_at >= week_start)
                .scalar()
                or 0
            )
            new_users_month = (
                session.query(func.count(User.id))
                .filter(User.created_at >= month_start)
                .scalar()
                or 0
            )

            # Sales (paid orders only)
            paid_filter = Order.status == "paid"

            def _sum(col, extra=None):
                q = session.query(func.coalesce(func.sum(col), 0))
                q = q.filter(paid_filter)
                if extra is not None:
                    q = q.filter(extra)
                return int(q.scalar() or 0)

            sales_today = _sum(Order.total_price_iqd, Order.created_at >= today_start)
            sales_week = _sum(Order.total_price_iqd, Order.created_at >= week_start)
            sales_month = _sum(Order.total_price_iqd, Order.created_at >= month_start)
            total_revenue = _sum(Order.total_price_iqd)
            total_cost = _sum(Order.cost_price_iqd)
            total_discount = _sum(Order.discount_amount)
            total_refunded = _sum(Order.refunded_amount)
            net_profit = total_revenue - total_cost - total_discount - total_refunded

            total_orders = (
                session.query(func.count(Order.id)).scalar() or 0
            )
            total_paid_orders = (
                session.query(func.count(Order.id))
                .filter(paid_filter)
                .scalar()
                or 0
            )

            # Top plans
            top_plans_rows = (
                session.query(
                    Plan.name,
                    func.count(Order.id).label("order_count"),
                    func.coalesce(func.sum(Order.total_price_iqd), 0).label("revenue"),
                )
                .join(Order, Order.plan_id == Plan.id)
                .filter(paid_filter)
                .group_by(Plan.name)
                .order_by(func.count(Order.id).desc())
                .limit(10)
                .all()
            )

            top_plans = [
                {
                    "name": r.name,
                    "order_count": r.order_count,
                    "revenue": int(r.revenue),
                }
                for r in top_plans_rows
            ]

            # Recent orders (all statuses)
            recent = (
                session.query(Order)
                .options(joinedload(Order.user), joinedload(Order.plan))
                .order_by(Order.created_at.desc())
                .limit(10)
                .all()
            )

            recent_orders = [
                {
                    "id": str(o.id),
                    "user_name": o.user.name if o.user else None,
                    "user_phone": o.user.phone if o.user else None,
                    "plan_name": o.plan.name if o.plan else None,
                    "total_price_iqd": o.total_price_iqd,
                    "status": o.status,
                    "created_at": o.created_at.isoformat(),
                }
                for o in recent
            ]

            return {
                "users": {
                    "total": total_users,
                    "active": active_users,
                    "new_today": new_users_today,
                    "new_week": new_users_week,
                    "new_month": new_users_month,
                },
                "orders": {
                    "total": total_orders,
                    "paid": total_paid_orders,
                },
                "sales": {
                    "today": sales_today,
                    "week": sales_week,
                    "month": sales_month,
                },
                "financial": {
                    "total_revenue": total_revenue,
                    "total_cost": total_cost,
                    "total_discount": total_discount,
                    "total_refunded": total_refunded,
                    "net_profit": net_profit,
                },
                "top_plans": top_plans,
                "recent_orders": recent_orders,
            }

    @staticmethod
    def get_sales_chart(days: int = 30) -> list[dict]:
        if days < 1 or days > 365:
            raise AppError(ErrorCode.VALIDATION_INVALID_PARAMETER)
        now = datetime.now(timezone.utc)
        since = now - timedelta(days=days)
        with get_session() as session:
            rows = (
                session.query(
                    func.date(Order.created_at).label("date"),
                    func.coalesce(func.sum(Order.total_price_iqd), 0).label("revenue"),
                    func.count(Order.id).label("order_count"),
                )
                .filter(
                    Order.status == "paid",
                    Order.created_at >= since,
                )
                .group_by(func.date(Order.created_at))
                .order_by(func.date(Order.created_at))
                .all()
            )

            data_map = {
                str(r.date): {
                    "revenue": int(r.revenue),
                    "order_count": r.order_count,
                }
                for r in rows
            }

            result = []
            for i in range(days):
                day = (since + timedelta(days=i + 1)).date()
                key = str(day)
                result.append({
                    "date": key,
                    "revenue": data_map.get(key, {}).get("revenue", 0),
                    "order_count": data_map.get(key, {}).get("order_count", 0),
                })
            return result

    @staticmethod
    def get_plan_distribution() -> list[dict]:
        with get_session() as session:
            rows = (
                session.query(
                    Plan.name,
                    func.count(Order.id).label("order_count"),
                    func.coalesce(func.sum(Order.total_price_iqd), 0).label("revenue"),
                )
                .join(Order, Order.plan_id == Plan.id)
                .filter(Order.status == "paid")
                .group_by(Plan.name)
                .order_by(func.count(Order.id).desc())
                .all()
            )
            return [
                {
                    "name": r.name,
                    "order_count": r.order_count,
                    "revenue": int(r.revenue),
                }
                for r in rows
            ]

    @staticmethod
    def get_user_growth(period: str = "daily") -> list[dict]:
        if period not in ("daily", "weekly", "monthly"):
            raise AppError(ErrorCode.VALIDATION_INVALID_PARAMETER)
        with get_session() as session:
            if period == "monthly":
                trunc = func.date_trunc("month", User.created_at)
            elif period == "weekly":
                trunc = func.date_trunc("week", User.created_at)
            else:
                trunc = func.date(User.created_at)
            rows = (
                session.query(
                    trunc.label("period"),
                    func.count(User.id).label("count"),
                )
                .group_by(trunc)
                .order_by(trunc.asc())
                .all()
            )
            return [
                {"period": str(r.period), "count": r.count}
                for r in rows
            ]

    @staticmethod
    def get_admin_activity(
        page: int = 1,
        limit: int = 20,
        action: str | None = None,
    ) -> dict:
        with get_session() as session:
            query = session.query(AuditLog).options(joinedload(AuditLog.user))
            query = query.filter(AuditLog.action.like("admin.%"))
            if action:
                query = query.filter(AuditLog.action == action)
            query = query.order_by(AuditLog.created_at.desc())
            total = query.count()
            offset = (page - 1) * limit
            items = query.offset(offset).limit(limit).all()
            return {
                "items": [
                    {
                        "id": str(item.id),
                        "user_id": str(item.user_id) if item.user_id else None,
                        "user_name": item.user.name if item.user else "",
                        "action": item.action,
                        "resource_type": item.resource_type,
                        "resource_id": item.resource_id,
                        "details": item.details,
                        "ip_address": item.ip_address,
                        "created_at": item.created_at.isoformat(),
                    }
                    for item in items
                ],
                "total": total,
                "page": page,
                "limit": limit,
            }

    # ── Report Export ─────────────────────────────────────────────

    @staticmethod
    def export_report(
        report_type: str,
        export_format: str,
        **filters: str | None,
    ) -> tuple[bytes | str, str, str]:
        if report_type not in _REPORT_TYPES:
            raise AppError(ErrorCode.VALIDATION_INVALID_PARAMETER)
        if export_format not in _EXPORT_FORMATS:
            raise AppError(ErrorCode.VALIDATION_INVALID_PARAMETER)

        data = AnalyticsService._fetch_report_data(report_type, **filters)
        now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{report_type}_report_{now_str}"

        if export_format == "csv":
            content = AnalyticsService._export_csv(data)
            mime = "text/csv"
            filename += ".csv"
        elif export_format == "pdf":
            content = AnalyticsService._export_pdf(report_type, data)
            mime = "application/pdf"
            filename += ".pdf"
        else:
            content = AnalyticsService._export_xlsx(report_type, data)
            mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename += ".xlsx"

        return content, mime, filename

    @staticmethod
    def _fetch_report_data(report_type: str, **filters: str | None) -> list[dict]:
        if report_type == "users":
            with get_session() as session:
                users = (
                    session.query(User)
                    .order_by(User.created_at.desc())
                    .all()
                )
                return [
                    {
                        "id": str(u.id),
                        "phone": u.phone,
                        "name": u.name,
                        "role": u.role,
                        "is_active": u.is_active,
                        "is_verified": u.is_verified,
                        "created_at": u.created_at.isoformat(),
                    }
                    for u in users
                ]

        if report_type == "orders":
            with get_session() as session:
                query = session.query(Order).options(
                    joinedload(Order.user),
                    joinedload(Order.plan),
                )
                status = filters.get("status")
                if status:
                    query = query.filter(Order.status == status)
                orders = query.order_by(Order.created_at.desc()).all()
                return [
                    {
                        "id": str(o.id),
                        "user_name": o.user.name if o.user else "",
                        "user_phone": o.user.phone if o.user else "",
                        "plan_name": o.plan.name if o.plan else "",
                        "quantity": o.quantity,
                        "total_price_iqd": o.total_price_iqd,
                        "status": o.status,
                        "cost_price_iqd": o.cost_price_iqd,
                        "refunded_amount": o.refunded_amount,
                        "created_at": o.created_at.isoformat(),
                    }
                    for o in orders
                ]

        if report_type == "payments":
            with get_session() as session:
                query = session.query(Payment).options(joinedload(Payment.user))
                status = filters.get("status")
                if status:
                    query = query.filter(Payment.status == status)
                method = filters.get("method")
                if method:
                    query = query.filter(Payment.method == method)
                payments = query.order_by(Payment.created_at.desc()).all()
                return [
                    {
                        "id": str(p.id),
                        "user_name": p.user.name if p.user else "",
                        "user_phone": p.user.phone if p.user else "",
                        "amount": p.amount,
                        "method": p.method,
                        "status": p.status,
                        "order_id": str(p.order_id) if p.order_id else None,
                        "created_at": p.created_at.isoformat(),
                    }
                    for p in payments
                ]

        # financial report
        stats = AnalyticsService.get_dashboard_stats()
        f = stats["financial"]
        s = stats["sales"]
        u = stats["users"]
        return [{
            "total_users": u["total"],
            "active_users": u["active"],
            "total_orders": stats["orders"]["total"],
            "paid_orders": stats["orders"]["paid"],
            "total_revenue": f["total_revenue"],
            "total_cost": f["total_cost"],
            "total_discount": f["total_discount"],
            "total_refunded": f["total_refunded"],
            "net_profit": f["net_profit"],
            "sales_today": s["today"],
            "sales_week": s["week"],
            "sales_month": s["month"],
        }]

    @staticmethod
    def _export_csv(data: list[dict]) -> str:
        output = StringIO()
        if not data:
            return output.getvalue()
        writer = csv.writer(output)
        writer.writerow(list(data[0].keys()))
        for row in data:
            writer.writerow(list(row.values()))
        return output.getvalue()

    @staticmethod
    def _export_pdf(report_type: str, data: list[dict]) -> bytes:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 12, f"{report_type.title()} Report", align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "I", 8)
        pdf.cell(0, 6, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(8)

        if not data:
            pdf.set_font("Helvetica", "", 10)
            pdf.cell(0, 8, "No data available.", new_x="LMARGIN", new_y="NEXT")
            return pdf.output()

        headers = list(data[0].keys())
        col_w = max(12, min(60, 180 // len(headers)))

        pdf.set_font("Helvetica", "B", 8)
        for h in headers:
            pdf.cell(col_w, 7, h.replace("_", " ").title(), border=1)
        pdf.ln()

        pdf.set_font("Helvetica", "", 7)
        for row in data:
            for h in headers:
                val = str(row.get(h, ""))
                pdf.cell(col_w, 6, val[:30], border=1)
            pdf.ln()
            if pdf.get_y() > 260:
                pdf.add_page()
                pdf.set_font("Helvetica", "B", 8)
                for h in headers:
                    pdf.cell(col_w, 7, h.replace("_", " ").title(), border=1)
                pdf.ln()
                pdf.set_font("Helvetica", "", 7)

        pdf.ln(10)
        pdf.set_font("Helvetica", "I", 8)
        pdf.cell(0, 5, f"Total rows: {len(data)}", align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 5, "ESIM EGO - Analytics Report", align="C")
        return pdf.output()

    @staticmethod
    def _export_xlsx(data: list[dict]) -> bytes:
        try:
            from openpyxl import Workbook
        except ImportError:
            raise AppError(ErrorCode.INTERNAL_ERROR)
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        if data:
            ws.append(list(data[0].keys()))
            for row in data:
                ws.append(list(row.values()))
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
